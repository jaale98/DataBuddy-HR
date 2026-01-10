from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from uuid import uuid4

from openpyxl import load_workbook

from app.ingest.schema import CANONICAL_COLUMNS, normalize_header


@dataclass
class IngestError(Exception):
    status_code: int
    error: str
    message: str
    details: dict


def ingest_file(
    original_path: Path, working_path: Path, max_rows: int
) -> dict[str, object]:
    suffix = original_path.suffix.lower()
    if suffix == ".csv":
        return _ingest_csv(original_path, working_path, max_rows)
    if suffix == ".xlsx":
        return _ingest_xlsx(original_path, working_path, max_rows)
    raise IngestError(
        status_code=400,
        error="unsupported_file",
        message="Upload rejected: unsupported file type.",
        details={},
    )


def _ingest_csv(original_path: Path, working_path: Path, max_rows: int) -> dict[str, object]:
    with original_path.open("r", newline="", encoding="utf-8") as in_file:
        reader = csv.reader(in_file)
        header = next(reader, None)
        if not header:
            raise IngestError(
                status_code=422,
                error="upload_rejected",
                message="Upload rejected: empty file.",
                details={"reason": "empty_file"},
            )

        column_map, unknown_columns, canonical_columns = _build_column_map(header)
        row_count = _write_rows(reader, working_path, column_map, canonical_columns, max_rows)

    return _dataset_meta(row_count, canonical_columns, unknown_columns)


def _ingest_xlsx(original_path: Path, working_path: Path, max_rows: int) -> dict[str, object]:
    try:
        workbook = load_workbook(filename=original_path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive parse guard
        raise IngestError(
            status_code=422,
            error="upload_rejected",
            message="Upload rejected: failed to parse XLSX.",
            details={"reason": "parse_error"},
        ) from exc

    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise IngestError(
            status_code=422,
            error="upload_rejected",
            message="Upload rejected: empty file.",
            details={"reason": "empty_file"},
        )

    header = ["" if value is None else str(value) for value in header_row]
    column_map, unknown_columns, canonical_columns = _build_column_map(header)
    row_count = _write_rows(rows, working_path, column_map, canonical_columns, max_rows)
    workbook.close()
    return _dataset_meta(row_count, canonical_columns, unknown_columns)


def _build_column_map(
    header: list[str],
) -> tuple[dict[str, int], list[str], list[str]]:
    canonical_map = {normalize_header(name): name for name in CANONICAL_COLUMNS}
    used_canonical: set[str] = set()
    unknown_columns: list[str] = []
    column_map: dict[str, int] = {}

    for index, column in enumerate(header):
        normalized = normalize_header(column)
        canonical = canonical_map.get(normalized)
        if canonical and canonical not in used_canonical:
            column_map[canonical] = index
            used_canonical.add(canonical)
        else:
            unknown_columns.append(column)

    canonical_columns = [name for name in CANONICAL_COLUMNS if name in column_map]
    return column_map, unknown_columns, canonical_columns


def _write_rows(
    rows: Iterable[Iterable[object]],
    working_path: Path,
    column_map: dict[str, int],
    canonical_columns: list[str],
    max_rows: int,
) -> int:
    row_count = 0
    working_path.parent.mkdir(parents=True, exist_ok=True)
    with working_path.open("w", newline="", encoding="utf-8") as out_file:
        writer = csv.writer(out_file)
        writer.writerow(["row_id", *canonical_columns])
        for row in rows:
            row_count += 1
            if row_count > max_rows:
                raise IngestError(
                    status_code=422,
                    error="upload_rejected",
                    message="Upload rejected: dataset exceeds the 50,000 row limit.",
                    details={
                        "reason": "too_many_rows",
                        "max_rows": max_rows,
                        "received_rows": row_count,
                    },
                )
            values = _row_values(row, column_map, canonical_columns)
            writer.writerow([str(uuid4()), *values])
    return row_count


def _row_values(
    row: Iterable[object],
    column_map: dict[str, int],
    canonical_columns: list[str],
) -> list[str]:
    row_list = list(row)
    values: list[str] = []
    for column in canonical_columns:
        index = column_map[column]
        value = row_list[index] if index < len(row_list) else None
        values.append(_format_cell(value))
    return values


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _dataset_meta(
    row_count: int, canonical_columns: list[str], unknown_columns: list[str]
) -> dict[str, object]:
    return {
        "total_rows": row_count,
        "total_columns": len(canonical_columns),
        "canonical_columns": canonical_columns,
        "unknown_columns": unknown_columns,
        "row_id_column": "row_id",
    }
